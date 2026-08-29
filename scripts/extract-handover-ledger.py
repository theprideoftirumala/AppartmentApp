"""One-off extractor. Reads the handed-over xlsx and writes src/data/handoverLedger.js."""
import json
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\vijaykumar.raavi\Downloads\The Pride of Tirumala - I&E Summary.xlsx")
OUT = Path(__file__).resolve().parents[1] / "src" / "data" / "handoverLedger.js"

CATS = {
    31: "Cleaning",
    32: "Generator",
    33: "Lift Service",
    34: "Service charges",
    35: "Repairs",
    36: "Garbage",
    37: "Electricity",
    38: "Internet",
    39: "Watchman salary",
    40: "Water",
    41: "Pest Control",
    42: "Sundry",
}


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Summary"]
    months = []
    for col in range(4, 74):
        header = ws.cell(10, col).value
        if not header:
            continue
        rec = {
            "label": str(header).strip(),
            "carryIn": None if ws.cell(11, col).value is None else round(float(ws.cell(11, col).value), 2),
            "collection": round(float(ws.cell(23, col).value or ws.cell(27, col).value or 0), 2),
            "expenses": round(float(ws.cell(43, col).value or 0), 2),
            "surplus": round(float(ws.cell(46, col).value or 0), 2),
            "byCategory": {},
        }
        for row, name in CATS.items():
            value = ws.cell(row, col).value
            if value not in (None, "", "-"):
                rec["byCategory"][name] = round(float(value), 2)
        months.append(rec)

    collection = round(sum(row["collection"] for row in months), 2)
    expenses = round(sum(row["expenses"] for row in months), 2)
    last = months[-1]
    months_js = json.dumps(months, indent=2)

    js = f"""/**
 * Figures copied from the society I&E workbook handed over on 29 Aug 2026.
 * Source: The Pride of Tirumala - I&E Summary.xlsx
 * Flat owner names are intentionally omitted.
 */

export const HANDOVER_SOURCE = 'The Pride of Tirumala - I&E Summary.xlsx';

export const HANDOVER_PROPERTY = {{
  name: 'The Pride of Tirumala',
  area: 'Alkapur, Neknampur',
  address: 'PLNo 49&48&47, Road No 20, 500089',
  firstContribution: '2020-11-01',
  flatCount: 10,
}};

/** Cash physically handed over. Not the sheet running total. */
export const HANDOVER_CASH_SURPLUS = 612;

export const HANDOVER_META = {{
  date: '2026-08-29',
  lastClosedMonth: {json.dumps(last["label"])},
  lastClosedCollection: {last["collection"]},
  lastClosedExpenses: {last["expenses"]},
  lastClosedMonthSurplus: {last["surplus"]},
  lastClosedCarryIn: {last["carryIn"]},
  lifetimeCollection: {collection},
  lifetimeExpenses: {expenses},
  sheetComputedNet: {round(collection - expenses, 2)},
  detailedExpenseRows: 759,
  lastDetailedDate: '2026-08-28',
  lastDetailedAmount: 1100,
  lastDetailedMemo: 'water tanker',
}};

export const LATEST_RECURRING = {{
  monthlyMaintenancePerFlat: 3000,
  watchmanSalary: 8500,
  garbage: 1500,
  waterCharges: 1417,
  electricityAug26: 2008,
  generatorAug26: 2000,
}};

/** Phones from the Notes tab only. No UPI IDs were present in the workbook. */
export const HANDOVER_CONTACTS = [
  ['Lift / Elevator', 'Lift service', '', '8074839972', '', '', 'From Notes tab'],
  ['Electrical', 'Electrician', '', '9381238096', '', '', 'From Notes tab'],
  ['Plumbing', 'Plumber', '', '8144134773', '', '', 'From Notes tab'],
  ['Garbage', 'Garbage collection', '', '9704744892', '', '', 'From Notes tab'],
];

export const HANDOVER_PAYEES = [
  ['Watchman salary', 'Watchman', 'Watchman', '', '', '8500', "Latest monthly amount on August '26 Summary. Add UPI in this tab."],
  ['Garbage', 'Garbage', 'Garbage vendor', '9704744892', '', '1500', "Latest monthly amount on August '26 Summary."],
  ['Water tanker', 'Water Tanker', 'Water tanker', '', '', '1100', 'Last detailed log line dated 28 Aug 2026.'],
  ['Lift service', 'Lift / Elevator', 'Lift AMC', '8074839972', '', '', 'Phone from Notes tab.'],
  ['Electrician', 'Electrical', 'Electrician', '9381238096', '', '', 'Phone from Notes tab.'],
  ['Plumber', 'Plumbing', 'Plumber', '8144134773', '', '', 'Phone from Notes tab.'],
];

export const HANDOVER_NOTES = [
  ['Wifi', 'Id recorded on Notes tab', 'Tirumala 2021'],
  ['CC TV id', 'Label only on Notes tab', ''],
  ['Generator service', 'Label only on Notes tab', ''],
  ['Tanker booking', 'Label only on Notes tab', ''],
  ['Electricity Dept', 'Label only on Notes tab', ''],
  ['CC Cam', 'Label only on Notes tab', ''],
  ['Borewell activity', 'Contribution A / B / motor / borewell from Borewell Exp tab', 'A 70000, B 360000, motor 59000, borewell 417500, noted result -2500'],
  ['Motor repair (Oct tab)', 'Equal collection recorded on Motor repair oct tab', '1750 per flat, 17500 total'],
];

export const HANDOVER_MONTHS = {months_js};

export function handoverLifetimeTotals(rows = HANDOVER_MONTHS) {{
  const collection = rows.reduce((sum, row) => sum + Number(row.collection || 0), 0);
  const expenses = rows.reduce((sum, row) => sum + Number(row.expenses || 0), 0);
  return {{
    collection,
    expenses,
    net: Math.round((collection - expenses) * 100) / 100,
  }};
}}
"""
    OUT.write_text(js, encoding="utf-8")
    print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
